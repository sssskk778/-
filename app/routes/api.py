"""
Модуль API маршрутов Carrier Rating Platform.

Автор: Лосева Е.А.
Дата создания: ДД.ММ.ГГГГ
Последнее изменение: ДД.ММ.ГГГГ
Контакт: ekaterinaloseva91@gmail.com
"""
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Blueprint, jsonify, request, make_response
from app import db
from app.auth import login_required, admin_required, current_user
from app.models import Criterion, Carrier, RunResult, Dataset, ScenarioCriterion, Run, Scenario, Shipment
from app.services.dataset_service import DatasetService
from app.services.scenario_service import ScenarioService
from app.services.run_service import RunService
from app.services.swara import SwaraService

api_bp = Blueprint('api', __name__, url_prefix='/api')
datasets = DatasetService()
scenarios = ScenarioService()
runs = RunService()


@api_bp.get('/auth/me')
@login_required
def me():
    u = current_user()
    return jsonify({
        'id': u.id,
        'username': u.username,
        'full_name': u.full_name,
        'role': u.role
    })


@api_bp.get('/datasets')
@login_required
def dataset_list():
    rows = datasets.list_datasets()
    return jsonify([{
        'id': d.id,
        'name': d.name,
        'file_name': d.file_name,
        'description': d.description,
        'records_count': d.records_count,
        'created_at': d.created_at.isoformat() if d.created_at else None
    } for d in rows])


@api_bp.get('/datasets/<int:did>')
@login_required
def dataset_detail(did):
    ds = datasets.get_dataset(did)
    carriers_count = Carrier.query.count()
    shipments_count = Shipment.query.filter_by(dataset_id=did).count()
    return jsonify({
        'id': ds.id,
        'name': ds.name,
        'file_name': ds.file_name,
        'description': ds.description,
        'records_count': ds.records_count,
        'carriers_count': carriers_count,
        'shipments_count': shipments_count,
        'created_at': ds.created_at.isoformat() if ds.created_at else None
    })


@api_bp.post('/datasets/upload')
@login_required
def upload_dataset():
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не передан.'}), 400
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Поддерживаются только Excel-файлы (.xlsx, .xls)'}), 400
    name = request.form.get('name') or file.filename
    description = request.form.get('description') or ''
    skip_preprocess = request.form.get('skip_preprocess', 'false').lower() == 'true'
    try:
        result = datasets.import_excel(
            file_storage=file, name=name, description=description,
            skip_preprocess=skip_preprocess
        )
        ds = result['dataset']
        preprocess_report = result.get('preprocess_report')
        response = {
            'id': ds.id, 'name': ds.name, 'file_name': ds.file_name,
            'records_count': ds.records_count,
            'carriers_count': Carrier.query.count(),
            'shipments_count': Shipment.query.filter_by(dataset_id=ds.id).count(),
            'message': 'Датасет успешно загружен'
        }
        if preprocess_report:
            response['preprocess'] = {
                'total_carriers': preprocess_report['stats']['total_carriers'],
                'valid_carriers': preprocess_report['stats']['valid_carriers'],
                'empty_carriers': preprocess_report['stats'].get('empty_carriers', 0),
                'total_shipments': preprocess_report['stats']['total_shipments'],
                'valid_shipments': preprocess_report['stats']['valid_shipments'],
                'empty_shipments': preprocess_report['stats'].get('empty_shipments', 0),
                'total_errors': preprocess_report['total_errors'],
                'errors': preprocess_report['errors'][:20]
            }
        return jsonify(response), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@api_bp.delete('/datasets/<int:did>')
@login_required
def delete_dataset(did):
    try:
        datasets.delete_dataset(did)
        return jsonify({'status': 'ok', 'message': 'Датасет удален'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@api_bp.get('/carriers')
@login_required
def carriers():
    rows = Carrier.query.order_by(Carrier.company_name.asc()).all()
    return jsonify([{
        'carrier_id': c.carrier_id,
        'company_name': c.company_name,
        'fleet_type': c.fleet_type,
        'region': c.region
    } for c in rows])


@api_bp.get('/carriers/<int:cid>')
@login_required
def carrier_detail(cid):
    carrier = Carrier.query.get_or_404(cid)
    shipments = Shipment.query.filter_by(carrier_id=cid).all()
    total_shipments = len(shipments)
    delivered = sum(1 for s in shipments if s.status == 'Доставлено')
    cancelled = sum(1 for s in shipments if s.status == 'Отменено')
    no_show = sum(1 for s in shipments if s.status == 'Не приехал')
    avg_rating = db.session.query(db.func.avg(Shipment.client_rating)) \
        .filter(Shipment.carrier_id == cid, Shipment.client_rating.isnot(None)).scalar()
    total_price = db.session.query(db.func.sum(Shipment.price)) \
        .filter(Shipment.carrier_id == cid).scalar() or 0
    return jsonify({
        'carrier_id': carrier.carrier_id,
        'company_name': carrier.company_name,
        'fleet_type': carrier.fleet_type,
        'region': carrier.region,
        'stats': {
            'total_shipments': total_shipments,
            'delivered': delivered,
            'cancelled': cancelled,
            'no_show': no_show,
            'avg_rating': round(avg_rating, 2) if avg_rating else None,
            'total_price': total_price
        }
    })


@api_bp.get('/shipments')
@login_required
def shipments():
    carrier_id = request.args.get('carrier_id', type=int)
    dataset_id = request.args.get('dataset_id', type=int)
    status = request.args.get('status')
    query = Shipment.query
    if carrier_id:
        query = query.filter_by(carrier_id=carrier_id)
    if dataset_id:
        query = query.filter_by(dataset_id=dataset_id)
    if status:
        query = query.filter_by(status=status)
    rows = query.order_by(Shipment.pickup_window_start.desc()).limit(100).all()
    return jsonify([{
        'shipment_id': s.shipment_id,
        'carrier_id': s.carrier_id,
        'dataset_id': s.dataset_id,
        'pickup_window_start': s.pickup_window_start.isoformat() if s.pickup_window_start else None,
        'delivery_window_end': s.delivery_window_end.isoformat() if s.delivery_window_end else None,
        'actual_delivery_time': s.actual_delivery_time.isoformat() if s.actual_delivery_time else None,
        'client_rating': s.client_rating,
        'price': float(s.price) if s.price else None,
        'distance_km': float(s.distance_km) if s.distance_km else None,
        'status': s.status,
        'has_gps': s.has_gps,
        'has_pod': s.has_pod,
        'accident_severity': s.accident_severity,
        'carrier_fault': s.carrier_fault,
        'claim_type': s.claim_type
    } for s in rows])


@api_bp.get('/criteria')
@login_required
def criteria():
    rows = Criterion.query.order_by(Criterion.order_no.asc()).all()
    return jsonify([{
        'id': c.id,
        'code': c.code,
        'name': c.name,
        'kind': c.kind,
        'order_no': c.order_no
    } for c in rows])


@api_bp.get('/scenarios')
@login_required
def scenario_list():
    rows = scenarios.list_all()
    out = []
    for s in rows:
        selected = ScenarioCriterion.query.filter_by(scenario_id=s.id).order_by(ScenarioCriterion.order_no.asc()).all()
        criteria = [db.session.get(Criterion, x.criterion_id) for x in selected]
        out.append({
            'id': s.id,
            'name': s.name,
            'description': s.description,
            'method': s.method,
            'status': s.status,
            'criterion_ids': [c.id for c in criteria if c],
            'criteria': [{'id': c.id, 'code': c.code, 'name': c.name, 'kind': c.kind} for c in criteria if c]
        })
    return jsonify(out)


@api_bp.get('/scenarios/<int:sid>')
@login_required
def scenario_get(sid):
    s = scenarios.get(sid)
    selected = ScenarioCriterion.query.filter_by(scenario_id=s.id).order_by(ScenarioCriterion.order_no.asc()).all()
    criteria = [db.session.get(Criterion, x.criterion_id) for x in selected]
    return jsonify({
        'id': s.id,
        'name': s.name,
        'description': s.description,
        'method': s.method,
        'status': s.status,
        'criterion_ids': [c.id for c in criteria if c],
        'criteria': [{'id': c.id, 'code': c.code, 'name': c.name, 'kind': c.kind} for c in criteria if c],
        'swara_config': s.get_swara_config()
    })


@api_bp.post('/scenarios')
@admin_required
def scenario_create():
    payload = request.get_json(force=True)
    s = scenarios.create(payload, current_user().id)
    return jsonify({'id': s.id, 'name': s.name}), 201


@api_bp.put('/scenarios/<int:sid>')
@admin_required
def scenario_update(sid):
    payload = request.get_json(force=True)
    s = scenarios.update(sid, payload)
    return jsonify({'id': s.id, 'name': s.name})


@api_bp.delete('/scenarios/<int:sid>')
@admin_required
def scenario_delete(sid):
    scenarios.delete(sid)
    return jsonify({'status': 'deleted'})


@api_bp.put('/scenarios/<int:sid>/weights/swara')
@admin_required
def scenario_set_swara_weights(sid):
    payload = request.get_json(force=True)
    ranking = payload.get('ranking', [])
    s_values = payload.get('s_values', [])
    if not ranking or not s_values:
        return jsonify({'error': 'ranking и s_values обязательны'}), 400
    if len(s_values) != len(ranking) - 1:
        return jsonify({'error': f'Ожидается {len(ranking) - 1} значений s_values'}), 400
    if not SwaraService.validate_s_values(s_values):
        return jsonify({'error': 's_values должны быть >= 0'}), 400
    scenario = Scenario.query.get_or_404(sid)
    scenario.set_swara_config(ranking, s_values)
    db.session.commit()
    weights = SwaraService.compute(ranking, s_values)
    return jsonify({
        'id': scenario.id,
        'swara_config': scenario.get_swara_config(),
        'weights': weights
    })


@api_bp.get('/scenarios/<int:sid>/weights/swara')
@login_required
def scenario_get_swara_weights(sid):
    scenario = Scenario.query.get_or_404(sid)
    config = scenario.get_swara_config()
    if not config:
        return jsonify({'ranking': [], 's_values': [], 'weights': {}})
    weights = SwaraService.compute(config['ranking'], config['s_values'])
    return jsonify({
        'ranking': config['ranking'],
        's_values': config['s_values'],
        'weights': weights
    })


@api_bp.post('/scenarios/<int:sid>/weights/swara/preview')
@login_required
def scenario_preview_swara_weights(sid):
    payload = request.get_json(force=True)
    ranking = payload.get('ranking', [])
    s_values = payload.get('s_values', [])
    if not ranking or not s_values:
        return jsonify({'error': 'ranking и s_values обязательны'}), 400
    if len(s_values) != len(ranking) - 1:
        return jsonify({'error': f'Ожидается {len(ranking) - 1} значений s_values'}), 400
    weights = SwaraService.preview_weights(ranking, s_values)
    return jsonify({'weights': weights})


@api_bp.post('/scenarios/<int:sid>/run')
@admin_required
def scenario_run(sid):
    try:
        run = runs.execute(sid, current_user().id)
        return jsonify({'run_id': run.id, 'status': run.status}), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@api_bp.get('/scenarios/<int:sid>/runs')
@login_required
def scenario_runs(sid):
    rows = Run.query.filter_by(scenario_id=sid).order_by(Run.id.desc()).all()
    return jsonify([{
        'id': r.id,
        'status': r.status,
        'started_at': r.started_at.isoformat() if r.started_at else None,
        'finished_at': r.finished_at.isoformat() if r.finished_at else None
    } for r in rows])


@api_bp.get('/runs/<int:rid>')
@login_required
def run_detail(rid):
    run = runs.get_run(rid)
    results = runs.get_run_results(run.id)
    return jsonify({
        'run': {
            'id': run.id,
            'status': run.status,
            'started_at': run.started_at.isoformat() if run.started_at else None,
            'finished_at': run.finished_at.isoformat() if run.finished_at else None
        },
        'meta': json.loads(run.meta_json or '{}'),
        'results': [{
            'company_name': r.carrier.company_name,
            'carrier_id': r.carrier_id,
            'rank': r.rank,
            'score': r.score,
            'details': json.loads(r.details_json or '{}')
        } for r in results]
    })


@api_bp.get('/scenarios/<int:sid>/latest-results')
@login_required
def latest_results(sid):
    run, results = runs.latest_results(sid)
    if not run:
        return jsonify({'run': None, 'results': [], 'meta': None})
    meta = json.loads(run.meta_json or '{}')
    out = []
    for r in results:
        out.append({
            'rank': r.rank,
            'score': r.score,
            'company_name': r.carrier.company_name,
            'carrier_id': r.carrier_id,
            'details': json.loads(r.details_json or '{}')
        })
    return jsonify({
        'run': {
            'id': run.id,
            'status': run.status,
            'started_at': run.started_at.isoformat() if run.started_at else None,
            'finished_at': run.finished_at.isoformat() if run.finished_at else None
        },
        'results': out,
        'meta': meta
    })


@api_bp.get('/scenarios/<int:sid>/export')
@login_required
def export_excel(sid):
    run, results = runs.latest_results(sid)
    if not run:
        return jsonify({'error': 'Нет результатов'}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F2747", end_color="0F2747", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Место', 'ID перевозчика', 'Название', 'Оценка']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, r in enumerate(results, 2):
        values = [r.rank, r.carrier_id, r.carrier.company_name, round(r.score, 4)]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=scenario_{sid}_results.xlsx'
    return response