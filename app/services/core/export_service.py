"""
Модуль сервиса экспорта данных.

Автор: Лосева Е.А.
Дата создания: ДД.ММ.ГГГГ
Последнее изменение: ДД.ММ.ГГГГ
Контакт: ekaterinaloseva91@gmail.com
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:

    def build_excel_results(self, results: list) -> bytes:
        """
        Назначение:
            Формирует Excel-файл с результатами запуска.
        Параметры:
            results (list[RunResult]): Список результатов.
        Возвращает:
            bytes: Содержимое xlsx-файла.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Результаты'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='0F2747', end_color='0F2747', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = ['Место', 'ID перевозчика', 'Название', 'Оценка']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        for row_idx, r in enumerate(results, 2):
            for col, value in enumerate([r.rank, r.carrier_id, r.carrier.company_name, round(r.score, 4)], 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = center
                cell.border = thin

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()